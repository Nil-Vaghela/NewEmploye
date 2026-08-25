from flask import Flask, render_template, request, redirect, url_for
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user
from flask import Flask, render_template, request, redirect, url_for, flash
from datetime import datetime as dt
import pytz
import pandas as pd
import datetime
import openpyxl
import os
import glob
import random
import schedule
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import logging
#import function
import Function
from multiprocessing import Process
from email.mime.text import MIMEText
import numpy as np
from dotenv import load_dotenv
load_dotenv()
os.environ['TZ'] = 'US/Eastern'
app = Flask(__name__)
# app.config['SERVER_NAME'] = 'www.gsucfa.com'
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'nilvaghela')
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Webpage template
HTML_TEMPLATE = Function.html_home_function()

# Excel file
today = datetime.date.today()
class Admin(UserMixin):
    id = "admin"
    password = "32156"

@login_manager.user_loader
def load_user(user_id):
    if user_id == "admin":
        return Admin()
    return None

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    return Function.LoginCode()

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

def update_points_file(user_id, username):
    points_file = "points.xlsx"

    # Check if the points file exists
    if os.path.isfile(points_file):
        df_points = pd.read_excel(points_file)
    else:
        # Create a new DataFrame if the file does not exist
        columns = ['Name', 'User ID', 'Points']
        df_points = pd.DataFrame(columns=columns)

    # Find the row corresponding to the user
    user_row = df_points[df_points['User ID'] == user_id]

    if user_row.empty:
        # If the user is not in the points file, add a new row
        new_row = {'Name': username, 'User ID': user_id, 'Points': 0}
        df_points = pd.concat([df_points, pd.DataFrame([new_row])], ignore_index=True)
    else:
        # If the user exists, update their points (this example keeps the current points unchanged)
        old_points = user_row['Points'].values[0]
        new_points = old_points  # Since averaging old_points with itself is unnecessary
        df_points.loc[user_row.index, 'Points'] = new_points

    # Save the updated DataFrame back to the Excel file
    df_points.to_excel(points_file, index=False)

@app.route('/users_list', methods=['GET'])
def users_list_view():
    return users_list()

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    df = pd.read_excel('points.xlsx')
    df = df[df['User ID'] != user_id]
    df.to_excel('points.xlsx', index=False)
    return redirect(url_for('users_list_view'))

def users_list():
    df = pd.read_excel('points.xlsx')
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        new_points = request.form.get('new_points')
        # Find the user and update their points
        df.loc[df['User ID'] == int(user_id), 'Points'] = int(new_points)
        # Save the DataFrame back to the file
        df.to_excel('points.xlsx', index=False)
        return redirect(url_for('home'))

    # Convert the DataFrame to a list of dictionaries for easy access in the template
    users = df.to_dict(orient='records')
    return render_template('user_list.html', users=users)
    # df = pd.read_excel("Employee Names.xlsx")
    # points_file = "Points.xlsx"
    # if os.path.isfile(points_file):
    #     df_points = pd.read_excel(points_file)
    #     df = df.merge(df_points, on=['User ID', 'Name'], how='left')
    # else:
    #     df['Points'] = "0"
    # users = df.to_dict('records')
    # return render_template('user_list.html', users=users)

from flask import session

@app.route('/edit_points/<int:user_id>', methods=['GET', 'POST'])
def edit_points(user_id):
    points_file = "points.xlsx"
    df_points = pd.read_excel(points_file)
    user_row = df_points[df_points['User ID'] == user_id]

    if request.method == 'POST':
        new_points = float(request.form['new_points'])
        df_points.loc[user_row.index, 'Points'] = new_points
        df_points.to_excel(points_file, index=False)
        return redirect(url_for('users_list_view'))

    df_employee = pd.read_excel('Employee Names.xlsx')
    employee_match = df_employee.loc[df_employee['User ID'] == int(user_id)]
    if employee_match.empty:
        return "User ID not found in the system.", 404
    employee_name = employee_match['Name'].values[0]
    rating = employee_match['Rating'].values[0]
   
    session['user_id'] = user_id # store the user id in session
    session['user_name'] = employee_name # store the user name in session

    files = glob.glob('2024*.xlsx')  # Get all Excel files starting with '2024'
    files = sorted(files)
    user_data = {}
    for file in files:
        df = pd.read_excel(file)
        user_df = df[df['Name'] == employee_name]
        user_data[file] = user_df.to_dict(orient='records')

    user_points = user_row['Points'].values[0]

    # check if user is a Student Leader
    is_leader = df_employee.loc[df_employee['Name'] == employee_name, 'Student Postion'].values[0] == 'Student Leader'
    session['is_leader'] = bool(is_leader)  # store the leader status in session

    return render_template('edit_point.html', user_id=user_id, user_points=user_points,user_data=user_data,username = employee_name, is_leader=is_leader,rating=rating)


@app.route('/rate_student_form/<int:user_id>', methods=['GET', 'POST'])
def rate_student_form(user_id):
    if request.method == 'POST':
        # Process form submission
        rating = request.form['rating']
        # Save the rating to the database
        # ...
        return redirect(url_for('rate_students'))

    # Show the rating form
    return render_template('rate_student_form.html', user_id=user_id)


def load_users_data_from_excel(file_name="Employee Names.xlsx"):
    df = pd.read_excel(file_name)
    users_data = df.set_index("Name")["User ID"].to_dict()
    return users_data

users_data = load_users_data_from_excel()

def get_eastern_time():
    eastern_tz = pytz.timezone('US/Eastern')
    eastern_time = dt.now(eastern_tz)
    return eastern_time

def get_excel_file_name():
    today = datetime.date.today()
    return f'{today}.xlsx'

@app.route('/attendance', methods=['GET', 'POST'])
@login_required
def index():
    EXCEL_FILE = get_excel_file_name() 
    if request.method == 'POST': 
        users_data = load_users_data_from_excel()  # Add this line to refresh user data
        user_id = int(request.form['username'])
        action = request.form['action']
        
        timestamp = get_eastern_time().strftime("%I:%M %p")

        if user_id in users_data.values():
            username = [name for name, id in users_data.items() if id == user_id][0]
        else:
            return "User ID not found in the system. Please enter valid user id"
        
        existing_data = pd.DataFrame(columns=['Name', 'Clock In', 'Clock Out', 'Lunch End', 'Lunch Start', 'Total Working Hours'])

        try:
            existing_data = pd.read_excel(EXCEL_FILE, index_col=None, engine='openpyxl')
            existing_data = existing_data[['Name', 'Clock In', 'Lunch Start', 'Lunch End', 'Clock Out', 'Total Working Hours']]

        except FileNotFoundError:
            existing_data = pd.DataFrame(columns=['Name', 'Clock In', 'Lunch Start', 'Lunch End', 'Clock Out', 'Total Working Hours'])

        new_data = existing_data.copy()

        if action == 'Clock In':
            already_clocked_in = (existing_data['Name'] == username) & (existing_data['Clock In'].notnull()) & (existing_data['Clock Out'].isnull())
            if already_clocked_in.any():
                return "You already Clocked In."
            else:
               # update_points_file(user_id, username)
                new_row = {'Name': username, 'Clock In': timestamp}
                new_data = pd.concat([new_data, pd.DataFrame(new_row, index=[0])], ignore_index=True)

        else:
            username_exists = username in existing_data['Name'].values

            if not username_exists:
                return "Username not found in the system. Please Clock In first."

            if action == 'Clock Out':
                available_rows = existing_data[(existing_data['Name'] == username) & (existing_data['Clock In'].notnull()) & (existing_data['Clock Out'].isnull())]
                if not available_rows.empty:
                    row_index = available_rows.index[-1]
                    clock_in = datetime.datetime.strptime(existing_data.loc[row_index, 'Clock In'], "%I:%M %p")  # Updated time format
                    clock_out = datetime.datetime.strptime(timestamp, "%I:%M %p")  # Updated time format
                    lunch_in = existing_data.loc[row_index, 'Lunch End']
                    lunch_out = existing_data.loc[row_index, 'Lunch Start']
                    if pd.notna(lunch_in) and pd.notna(lunch_out):
                        lunch_in = datetime.datetime.strptime(lunch_in, "%I:%M %p")
                        lunch_out = datetime.datetime.strptime(lunch_out, "%I:%M %p")
                        lunch_break = (lunch_in - lunch_out).total_seconds() / 3600
                        total_hours = (clock_out - clock_in).total_seconds() / 3600 - lunch_break
                        perfect_hours, minutes = divmod(int(total_hours * 60), 60)
                        total_hours = f'{perfect_hours}:{minutes}'
                    else:
                        total_hours = (clock_out - clock_in).total_seconds() / 3600
                        perfect_hours, minutes = divmod(int(total_hours * 60), 60)
                        total_hours = f'{perfect_hours}:{minutes}'
                    new_data.loc[row_index, 'Clock Out'] = timestamp
                    new_data.loc[row_index, 'Total Working Hours'] = total_hours
            elif action == 'Lunch Start':
                available_rows = existing_data[(existing_data['Name'] == username) & (existing_data['Clock In'].notnull()) & (existing_data['Clock Out'].isnull()) & (existing_data['Lunch Start'].isnull()) & (existing_data['Lunch End'].isnull())]
                if not available_rows.empty:
                    row_index = available_rows.index[-1]
                    new_data.loc[row_index, 'Lunch Start'] = timestamp

            elif action == 'Lunch End':
                available_rows = existing_data[(existing_data['Name'] == username) & (existing_data['Clock In'].notnull()) & (existing_data['Clock Out'].isnull()) & (existing_data['Lunch Start'].notnull()) & (existing_data['Lunch End'].isnull())]
                if not available_rows.empty:
                    row_index = available_rows.index[-1]
                    new_data.loc[row_index, 'Lunch End'] = timestamp

        with pd.ExcelWriter(EXCEL_FILE, mode='w', engine='openpyxl') as writer:
            new_data.to_excel(writer, index=False)

        return redirect('/attendance')

    
    try:
        existing_data = pd.read_excel(EXCEL_FILE, index_col=None, engine='openpyxl')
        emojis = ['🐈', '🦈', '👻 ', '🐶', '🦁', '🐬', '🐳', '🦕', '🦖','🐕','🐘','🦧','🦋','🐥','🦄','🙉','🐼','🫶','👑','🧡','💙','🐻‍❄️','🐊','🦘','🎷','🧸','🥹','🧁','🐅','🐰','🐝','🌵','🌸','🌝','☃️','🫠','😚','🤪','🤯','🥱','🐲','🦙','🪺','🎤']
        existing_data['    '] = [random.choice(emojis) for _ in existing_data.index]
        cols = existing_data.columns.tolist()
        cols = cols[-1:] + cols[:-1]
        existing_data = existing_data[cols]
        table_html = existing_data.to_html(classes=['table', 'table-striped'], index=False, justify='center')
    except FileNotFoundError:
        table_html = "<p>No data available.</p>"

    return HTML_TEMPLATE.format(table=table_html)

folder_path = './'
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        user_id = request.form['username']
        password = request.form['password']
        if user_id == Admin.id and password == "qwerty":
            login_user(Admin(), remember=True, duration=datetime.timedelta(hours=15))
            files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
            files = sorted(files)
            return render_template('admin.html',files=files)
        else:
            return "Invalid credentials. Please try again."

    return Function.adminLogin()

# @app.route('/admin')
# def admin():
#     files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
#     return render_template('admin.html',files=files)

@app.route('/view/<filename>')
def view_file(filename):
    file_path = os.path.join(folder_path, filename)
    df = pd.read_excel(file_path)
    return render_template('view_file.html', filename=filename, data=df.to_html(index=False, classes='table'))

@app.route('/edit/<filename>', methods=['GET', 'POST'])
def edit_file(filename):
    file_path = os.path.join(folder_path, filename)
    if request.method == 'POST':
        data = request.form.to_dict(flat=False)
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)
        return redirect(url_for('view_file', filename=filename))
    else:
        df = pd.read_excel(file_path)
        return render_template('edit_file.html', filename=filename, data=df.to_dict('records'))



def send_email_with_attachment():
    sender_email = os.environ.get("EMAIL_ADDRESS", "nilrajsinh6499@gmail.com")
    receiver_emails = ["tsmith327@gsu.edu", "lmorgan28@gsu.edu", "jjohnson626@student.gsu.edu", "nvaghela2@student.gsu.edu"]
    receiver_emails = ["nvaghela2@student.gsu.edu"]
    password = os.environ.get("EMAIL_PASSWORD", "vdbozwhhbcoynfby")

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = f"Attendance Report for {today}.xlsx"

    # Add body to the email
    body = "Dear all,\n\nPlease find the attached attendance report for today.\n\nBest regards,\nNil Vaghela ❤️\n \n"
    msg.attach(MIMEText(body, 'plain'))

    file_path = f"{today}.xlsx"
    attachment = open(file_path, "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f"attachment; filename={today}.xlsx")

    msg.attach(part)

    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_emails, msg.as_string())
    server.quit() 

    print(f"Attendance report for {today} has been sent to {', '.join(receiver_emails)}")



@app.route('/send_email', methods=['GET'])
@login_required
def send_email():
    send_email_with_attachment()
    return "Email has been sent successfully."

@app.route('/add_user', methods=['GET', 'POST'])
@login_required
def add_user():
    if request.method == 'POST':
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        Kitchen_Place = request.form['Back/Front']
        new_user_id = add_user_to_excel(first_name, last_name,Kitchen_Place)
        update_points_file(username= f"{first_name} {last_name[0]}",user_id=new_user_id, )
        flash(f"User {first_name} {last_name[0]}. has been added with User ID: {new_user_id}", "success")
        return redirect(url_for('add_user'))
    return render_template('add_user.html')


def add_user_to_excel(first_name, last_name, Kitchen_place):
    file_name = "Employee Names.xlsx"
    
    # Read the existing Excel file
    df = pd.read_excel(file_name)
    
    # Determine the new User ID
    max_user_id = df['User ID'].max()
    new_user_id = max_user_id + 1 if pd.notna(max_user_id) else 1
    
    # Create a new row with the user information
    new_row = {
        'Name': f'{first_name} {last_name[0]}.',
        'User ID': new_user_id,
        'Kitchen Place': Kitchen_place,
        'Student Postion': "Student Assistance",
        'Rating': 0
    }
    
    # Add the new row to the DataFrame using pd.concat
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    
    # Write the updated DataFrame back to the Excel file
    df.to_excel(file_name, index=False)
    
    # Call the update_points_file function (assuming it's defined elsewhere)
    update_points_file(username=f"{first_name} {last_name[0]}", user_id=new_user_id)
    
    return new_user_id

@app.route('/leave_request', methods=['GET'])
@login_required
def leave_request():
    return render_template('leave_request.html')



def send_leave_request_email(user_id, leave_start, leave_end, reason):
    sender_email = os.environ.get("EMAIL_ADDRESS", "nilrajsinh6499@gmail.com")
    receiver_emails = ["tsmith327@gsu.edu", "lmorgan28@gsu.edu","nvaghela2@student.gsu.edu"]
    #receiver_emails = ["nvaghela2@student.gsu.edu"]
    password = os.environ.get("EMAIL_PASSWORD", "vdbozwhhbcoynfby")

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = f"Leave Request"
    body = f"Hello \n {user_id} has requsted a leave for following dates and reason. \n\n Start Date: {leave_start} \n\n End Date: {leave_end} \n\n Reason: {reason}\n\n"
    #body = f"User ID {user_id} has requested a leave from {leave_start} to {leave_end}.\n\nReason:\n{reason}\n\n"
    msg.attach(MIMEText(body, 'plain'))

    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_emails, msg.as_string())
    server.quit()
    flash("Leave request submitted successfully.", "success")
    print(f"Leave request email sent to {', '.join(receiver_emails)}")
    return redirect(url_for('leave_request'))


@app.route('/apply_leave', methods=['POST'])
@login_required
def apply_leave():
    user_id = request.form['username']
    leave_start = request.form['leave_start']
    leave_end = request.form['leave_end']
    reason = request.form['reason']
    send_leave_request_email(user_id, leave_start, leave_end, reason)
    flash("Leave request submitted successfully.", "success")
    return redirect(url_for('leave_request'))



@app.route('/Feedback', methods=['GET', 'POST'])
def Feedback():
    if request.method == 'POST':
        name = request.form.get('name', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone_number', '')
        feedback = request.form.get('feedback', '')
        rating = request.form.get('rating', '')

        send_feedback_email(name, email, phone, feedback, rating)
        flash('Your feedback has been submitted', 'success')
    return render_template('Feedback.html')

def send_feedback_email(name, email, phone, feedback, rating):
    sender_email = os.environ.get("EMAIL_ADDRESS", "nilrajsinh6499@gmail.com")
    receiver_emails = ["tsmith327@gsu.edu", "lmorgan28@gsu.edu","nvaghela2@student.gsu.edu"]
    password = os.environ.get("EMAIL_PASSWORD", "vdbozwhhbcoynfby")

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = f"New Feedback Received"

    body = f"Hello,\n\nYou have received a new feedback.\n\nName: {name}\nEmail: {email}\nPhone Number: {phone}\nFeedback: {feedback}\nRating: {rating}\n\nBest,\nYour Feedback Team"
    msg.attach(MIMEText(body, 'plain'))

    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_emails, msg.as_string())
    server.quit()
    
    print(f"Feedback email sent to {', '.join(receiver_emails)}")

@app.route('/delete/<filename>/<int:index>', methods=['GET'])
@login_required
def delete_row(filename, index):
    file_path = os.path.join(folder_path, filename)
    df = pd.read_excel(file_path)
    df = df.drop(df.index[index])
    df.to_excel(file_path, index=False)
    return redirect(url_for('edit_file', filename=filename))


# if __name__ == '__main__':
#     app.run(host='0.0.0.0', port=80)




from itertools import cycle

from random import shuffle

@app.route('/assign', methods=['GET'])
def assign_students():
    df = pd.read_excel('Employee Names.xlsx')

    # Split the data into two dataframes based on the 'Student Postion' field
    student_assistance = df[df['Student Postion'] == 'Student Assistance'].copy()
    student_leader = df[df['Student Postion'] == 'Student Leader'].copy()

    # Shuffle the student leaders and assign them to student assistants
    leaders_list = list(student_leader['User ID'])
    if not leaders_list:
        return "No Student Leaders found. Add at least one before assigning.", 400
    if student_assistance.empty:
        return "No Student Assistants found. Add at least one before assigning.", 400
    shuffle(leaders_list)
    repeated_leaders = leaders_list * (len(student_assistance) // len(leaders_list)) + leaders_list[:len(student_assistance) % len(leaders_list)]
    student_assistance['Student Leader'] = repeated_leaders

    # Assign student leaders to other leaders, ensuring no self-assignment
    shuffled_leaders = leaders_list.copy()
    shuffle(shuffled_leaders)
    if len(leaders_list) > 1:
        for i, leader in enumerate(leaders_list):
            if leader == shuffled_leaders[i]:
                next_index = (i + 1) % len(leaders_list)
                shuffled_leaders[i], shuffled_leaders[next_index] = shuffled_leaders[next_index], shuffled_leaders[i]

    student_leader['Lead Student Leader'] = shuffled_leaders

    # Save both dataframes to new excel files
    student_assistance.to_excel('student_assistance_and_leaders.xlsx', index=False)
    student_leader.to_excel('student_leaders_to_leaders.xlsx', index=False)

    # Return a success message or redirect
    return "Student assistants and leaders have been assigned successfully!", 200


from openpyxl import load_workbook

@app.route('/rate_students', methods=['GET', 'POST'])
def rate_students():
    # Needs an active session (set by visiting /edit_points/<user_id> first)
    if 'user_id' not in session:
        return redirect(url_for('login'))
    # check if user is a Student Leader
    if 'is_leader' in session and session['is_leader']:
        # if the user is a Student Leader, find their assigned students
        df_students = pd.read_excel('student_assistance_and_leaders.xlsx')
        assigned_students = df_students[df_students['Student Leader'].astype(str) == str(session.get('user_id', ''))]
        if request.method == 'POST':
            user_id = request.form.get('user_id')
            total = sum(float(request.form.get(f'q{i+1}')) for i in range(10))
            avg_rating = total / 10
            score = avg_rating * 20

            # Update the Excel file with the student's average rating
            wb = load_workbook(filename='Employee Names.xlsx')
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[1] == str(user_id):  # Compare with the student's user_id
                    ws.cell(row=idx, column=5, value=score)  # Updated this to correctly reference the 'Rating' column
                    wb.save('Employee Names.xlsx')
                    break

        return render_template('rate_students.html', students=assigned_students.to_dict(orient='records'), sesstionUser=str(session.get('user_id', '')))
    else:
        # if the user is not a Student Leader, show an error message
        return render_template('rate_students.html', error="Only for Student Leader", sesstionUser=str(session.get('user_id', '')))

@app.route('/leaderboard')
def leaderboard():
    # Read data from Employee name Excel file
    df = pd.read_excel('Employee Names.xlsx')
    # Convert dataframe to dictionary
    sorted_employees = df.sort_values(by='Rating', ascending=False).to_dict('records')


    return render_template('leaderboard.html', employees=sorted_employees, sesstionUser=str(session.get('user_id', '')))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=os.environ.get('FLASK_DEBUG', '0') == '1')
    